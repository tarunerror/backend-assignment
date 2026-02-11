import os

from celery import shared_task
from django.conf import settings
from django.core.management.color import no_style
from django.db import connection
from openpyxl import load_workbook


def _normalize_header(value):
    if value is None:
        return ""
    return str(value).strip().lower().replace("_", " ")


def _get_cell_value(row, header_map, aliases, default=None):
    for alias in aliases:
        idx = header_map.get(alias)
        if idx is not None and idx < len(row):
            return row[idx]
    return default


def _to_int(value, default=0):
    if value in (None, ""):
        return default
    return int(value)


def _to_float(value, default=0.0):
    if value in (None, ""):
        return default
    return float(value)


def _sync_model_sequence(model):
    statements = connection.ops.sequence_reset_sql(no_style(), [model])
    if not statements:
        return
    with connection.cursor() as cursor:
        for statement in statements:
            cursor.execute(statement)


@shared_task
def ingest_customer_data():
    """Ingest customer data from customer_data.xlsx using background worker."""
    from .models import Customer

    file_path = os.path.join(settings.BASE_DIR, "customer_data.xlsx")
    wb = load_workbook(file_path)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return "No customer data found"

    header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}

    customers = []
    for row in rows:
        if all(v is None for v in row):
            continue

        customer_id = _to_int(
            _get_cell_value(row, header_map, ["customer id", "customerid"]),
            default=None,
        )
        if customer_id is None:
            continue

        first_name = _get_cell_value(
            row, header_map, ["first name", "firstname"], default=""
        )
        last_name = _get_cell_value(
            row, header_map, ["last name", "lastname"], default=""
        )
        age = _get_cell_value(row, header_map, ["age"], default=None)
        phone_number = _to_int(
            _get_cell_value(row, header_map, ["phone number", "phonenumber"]), default=0
        )
        monthly_salary = _to_int(
            _get_cell_value(
                row, header_map, ["monthly salary", "monthly income"], default=0
            ),
            default=0,
        )
        approved_limit = _to_int(
            _get_cell_value(row, header_map, ["approved limit"], default=0), default=0
        )
        current_debt = _to_float(
            _get_cell_value(row, header_map, ["current debt"], default=0), default=0
        )

        customers.append(
            Customer(
                customer_id=customer_id,
                first_name=str(first_name or ""),
                last_name=str(last_name or ""),
                age=_to_int(age, default=0) if age is not None else None,
                phone_number=phone_number,
                monthly_salary=monthly_salary,
                approved_limit=approved_limit,
                current_debt=current_debt,
            )
        )

    Customer.objects.bulk_create(customers, ignore_conflicts=True)
    _sync_model_sequence(Customer)
    return f"Ingested {len(customers)} customers"


@shared_task
def ingest_loan_data():
    """Ingest loan data from loan_data.xlsx using background worker."""
    from .models import Customer, Loan

    file_path = os.path.join(settings.BASE_DIR, "loan_data.xlsx")
    wb = load_workbook(file_path)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return "No loan data found"

    header_map = {_normalize_header(col): idx for idx, col in enumerate(header)}

    existing_customer_ids = set(Customer.objects.values_list("customer_id", flat=True))

    loans = []
    for row in rows:
        if all(v is None for v in row):
            continue

        customer_id = _to_int(
            _get_cell_value(row, header_map, ["customer id", "customerid"]),
            default=None,
        )
        loan_id = _to_int(
            _get_cell_value(row, header_map, ["loan id", "loanid"]), default=None
        )
        loan_amount = _to_float(
            _get_cell_value(row, header_map, ["loan amount"], default=0), default=0
        )
        tenure = _to_int(_get_cell_value(row, header_map, ["tenure"], default=0))
        interest_rate = _to_float(
            _get_cell_value(row, header_map, ["interest rate"], default=0), default=0
        )
        monthly_payment = _to_float(
            _get_cell_value(
                row,
                header_map,
                ["monthly payment", "monthly repayment", "emi"],
                default=0,
            ),
            default=0,
        )
        emis_paid_on_time = _to_int(
            _get_cell_value(
                row,
                header_map,
                ["emis paid on time", "emis paid ontime", "emis_paid_on_time"],
                default=0,
            ),
            default=0,
        )
        date_of_approval = _get_cell_value(
            row, header_map, ["date of approval", "start date"], default=None
        )
        end_date = _get_cell_value(row, header_map, ["end date"], default=None)

        if customer_id is None or loan_id is None:
            continue
        if customer_id not in existing_customer_ids:
            continue
        if hasattr(date_of_approval, "date"):
            date_of_approval = date_of_approval.date()
        if hasattr(end_date, "date"):
            end_date = end_date.date()
        loans.append(
            Loan(
                loan_id=loan_id,
                customer_id=customer_id,
                loan_amount=loan_amount,
                tenure=tenure,
                interest_rate=interest_rate,
                monthly_repayment=monthly_payment,
                emis_paid_on_time=emis_paid_on_time,
                date_of_approval=date_of_approval,
                end_date=end_date,
            )
        )

    Loan.objects.bulk_create(loans, ignore_conflicts=True)
    _sync_model_sequence(Loan)

    # Update current_debt for each customer
    from datetime import date

    from django.db.models import Sum

    for cid in existing_customer_ids:
        current_loans = Loan.objects.filter(customer_id=cid, end_date__gte=date.today())
        total_debt = current_loans.aggregate(total=Sum("loan_amount"))["total"] or 0
        Customer.objects.filter(customer_id=cid).update(current_debt=total_debt)

    return f"Ingested {len(loans)} loans"
